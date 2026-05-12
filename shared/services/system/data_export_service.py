"""
数据导出服务
提供CSV、Excel等格式的数据导出功能
"""

import csv
import io
import logging
from datetime import datetime
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)


class DataExportService:
    """数据导出服务"""

    def __init__(self):
        pass

    def export_to_csv(self, data: List[Dict],
                      filename: str = None,
                      encoding: str = 'utf-8-sig') -> bytes:
        """
        导出数据为CSV格式
        
        Args:
            data: 数据列表(字典列表)
            filename: 文件名(可选)
            encoding: 编码格式(默认utf-8-sig支持Excel中文)
            
        Returns:
            CSV文件字节流
        """
        if not data:
            raise ValueError("Data is empty")

        # 获取所有字段名
        fieldnames = list(data[0].keys())

        # 创建CSV缓冲区
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)

        # 写入表头和数据
        writer.writeheader()
        for row in data:
            # 处理特殊类型
            processed_row = {}
            for key, value in row.items():
                if isinstance(value, datetime):
                    processed_row[key] = value.isoformat()
                elif isinstance(value, (dict, list)):
                    processed_row[key] = str(value)
                else:
                    processed_row[key] = value
            writer.writerow(processed_row)

        # 转换为字节流
        csv_bytes = output.getvalue().encode(encoding)
        output.close()

        logger.info(f"Exported {len(data)} rows to CSV")
        return csv_bytes

    def export_to_excel(self, data: List[Dict],
                        filename: str = None,
                        sheet_name: str = 'Sheet1') -> bytes:
        """
        导出数据为Excel格式
        
        Args:
            data: 数据列表(字典列表)
            filename: 文件名(可选)
            sheet_name: 工作表名称
            
        Returns:
            Excel文件字节流
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if not data:
            raise ValueError("Data is empty")

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 获取字段名
        fieldnames = list(data[0].keys())

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col_idx, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=self._format_header(fieldname))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, fieldname in enumerate(fieldnames, 1):
                value = row_data.get(fieldname)

                # 处理特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    value = '是' if value else '否'
                elif value is None:
                    value = ''

                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 自动调整列宽
                if len(str(value)) > 20:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        output.close()

        logger.info(f"Exported {len(data)} rows to Excel")
        return excel_bytes

    def _format_header(self, fieldname: str) -> str:
        """
        格式化表头名称
        
        Args:
            fieldname: 字段名
            
        Returns:
            格式化后的表头
        """
        # 将snake_case转换为更易读的格式
        return fieldname.replace('_', ' ').title()

    def export_user_list(self, users: List[Dict],
                         format: str = 'csv') -> bytes:
        """
        导出用户列表
        
        Args:
            users: 用户数据列表
            format: 导出格式(csv/excel)
            
        Returns:
            文件字节流
        """
        # 选择需要导出的字段
        export_fields = [
            'id', 'username', 'email', 'phone',
            'is_active', 'is_verified', 'created_at',
            'last_login', 'article_count', 'follower_count',
        ]

        # 过滤数据
        filtered_users = []
        for user in users:
            filtered_user = {
                field: user.get(field, '')
                for field in export_fields
            }
            filtered_users.append(filtered_user)

        if format == 'excel':
            return self.export_to_excel(filtered_users, sheet_name='Users')
        else:
            return self.export_to_csv(filtered_users)

    def export_articles(self, articles: List[Dict],
                        format: str = 'csv') -> bytes:
        """
        导出文章列表
        
        Args:
            articles: 文章数据列表
            format: 导出格式(csv/excel)
            
        Returns:
            文件字节流
        """
        # 选择需要导出的字段
        export_fields = [
            'id', 'title', 'author_id', 'category_id',
            'status', 'view_count', 'like_count',
            'comment_count', 'created_at', 'updated_at',
            'published_at',
        ]

        # 过滤数据
        filtered_articles = []
        for article in articles:
            filtered_article = {
                field: article.get(field, '')
                for field in export_fields
            }
            filtered_articles.append(filtered_article)

        if format == 'excel':
            return self.export_to_excel(filtered_articles, sheet_name='Articles')
        else:
            return self.export_to_csv(filtered_articles)

    def export_comments(self, comments: List[Dict],
                        format: str = 'csv') -> bytes:
        """
        导出评论列表
        
        Args:
            comments: 评论数据列表
            format: 导出格式(csv/excel)
            
        Returns:
            文件字节流
        """
        # 选择需要导出的字段
        export_fields = [
            'id', 'article_id', 'user_id', 'content',
            'parent_id', 'like_count', 'created_at',
            'is_approved',
        ]

        # 过滤数据
        filtered_comments = []
        for comment in comments:
            filtered_comment = {
                field: comment.get(field, '')
                for field in export_fields
            }
            filtered_comments.append(filtered_comment)

        if format == 'excel':
            return self.export_to_excel(filtered_comments, sheet_name='Comments')
        else:
            return self.export_to_csv(filtered_comments)

    def export_analytics(self, analytics_data: List[Dict],
                         format: str = 'csv',
                         sheet_name: str = 'Analytics') -> bytes:
        """
        导出分析数据
        
        Args:
            analytics_data: 分析数据列表
            format: 导出格式(csv/excel)
            sheet_name: 工作表名称
            
        Returns:
            文件字节流
        """
        if format == 'excel':
            return self.export_to_excel(analytics_data, sheet_name=sheet_name)
        else:
            return self.export_to_csv(analytics_data)

    def get_export_templates(self) -> Dict[str, List[str]]:
        """
        获取可用的导出模板
        
        Returns:
            模板字典 {template_name: [fields]}
        """
        return {
            'users': [
                'id', 'username', 'email', 'phone',
                'is_active', 'is_verified', 'created_at',
                'last_login', 'article_count', 'follower_count',
            ],
            'articles': [
                'id', 'title', 'author_id', 'category_id',
                'status', 'view_count', 'like_count',
                'comment_count', 'created_at', 'updated_at',
                'published_at',
            ],
            'comments': [
                'id', 'article_id', 'user_id', 'content',
                'parent_id', 'like_count', 'created_at',
                'is_approved',
            ],
        }


# 全局实例
data_export_service = DataExportService()
